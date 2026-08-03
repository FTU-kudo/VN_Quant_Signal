"""
Auto-updater cho live_signal_tracker.
Chạy hàng ngày sau run_daily.py để cập nhật
trạng thái WIN/LOSS/HOLDING dựa trên giá thực tế.
"""
import pandas as pd
import numpy as np
from datetime import date, timedelta
from pathlib import Path
from config.settings import PROC_DIR

TRACKER_PATH = PROC_DIR / 'live_signal_tracker.parquet'
TRACKER_CSV  = PROC_DIR / 'live_signal_tracker.csv'


def update_signal_status(df_tracker: pd.DataFrame,
                         df_prices: pd.DataFrame) -> pd.DataFrame:
    """
    Cập nhật trạng thái cho các signals đang HOLDING.
    
    Logic xác định WIN/LOSS:
    
    WIN  → giá hiện tại >= TARGET_1
           (đạt mục tiêu chốt lời +2×ATR)
    
    LOSS → giá hiện tại <= STOP_LOSS
           (chạm cắt lỗ -2×ATR)
    
    EXPIRED → đã giữ > 20 ngày giao dịch
              (hết holding period backtest)
              Tự động tính return tại giá hiện tại.
    
    HOLDING → chưa chạm ngưỡng nào, tiếp tục theo dõi.
    """
    df = df_tracker.copy()

    # Ensure required columns exist
    for col in ['result', 'current_price', 'return_pct', 'last_updated', 'days_held', 'exit_price', 'exit_date', 'note']:
        if col not in df.columns:
            df[col] = None

    # Chỉ update các signals chưa đóng (result is NA hoặc 'HOLDING')
    holding_mask = df['result'].isna() | (df['result'] == 'HOLDING')
    if not holding_mask.any():
        return df
    
    # Latest price cho mỗi ticker
    latest_prices = (
        df_prices.sort_values('time')
                 .groupby('ticker')['close']
                 .last()
                 .to_dict()
    )
    latest_date = df_prices['time'].max().date()
    
    for idx in df[holding_mask].index:
        row    = df.loc[idx]
        ticker = row['ticker']
        
        current_price = latest_prices.get(ticker)
        if current_price is None:
            continue
        
        # Cập nhật current price và return
        df.at[idx, 'current_price'] = current_price
        df.at[idx, 'return_pct'] = round(
            (current_price - row['entry_price'])
            / row['entry_price'] * 100, 2
        )
        df.at[idx, 'last_updated'] = latest_date
        
        # Tính số ngày giao dịch đã giữ
        signal_date = pd.Timestamp(row['signal_date']).date()
        trading_days_held = len(
            df_prices[
                (df_prices['ticker'] == ticker) &
                (df_prices['time'].dt.date > signal_date) &
                (df_prices['time'].dt.date <= latest_date)
            ]
        )
        df.at[idx, 'days_held'] = trading_days_held
        
        # Xác định trạng thái
        sl = row.get('STOP_LOSS', 0)
        t1 = row.get('TARGET_1', float('inf'))
        
        if current_price >= t1:
            df.at[idx, 'result']     = 'WIN'
            df.at[idx, 'status']     = 'WIN'
            df.at[idx, 'exit_price'] = current_price
            df.at[idx, 'exit_date']  = latest_date
            
        elif current_price <= sl and sl > 0:
            df.at[idx, 'result']     = 'LOSS'
            df.at[idx, 'status']     = 'LOSS'
            df.at[idx, 'exit_price'] = current_price
            df.at[idx, 'exit_date']  = latest_date
            
        elif trading_days_held >= 20:
            # Hết holding period → đóng lệnh theo giá hiện tại
            res = 'WIN' if current_price > row['entry_price'] else 'LOSS'
            df.at[idx, 'result']     = res
            df.at[idx, 'status']     = res
            df.at[idx, 'exit_price'] = current_price
            df.at[idx, 'exit_date']  = latest_date
            df.at[idx, 'note']       = 'EXPIRED_T+20'
        else:
            df.at[idx, 'result']     = 'HOLDING'
            df.at[idx, 'status']     = 'HOLDING'
    
    return df


def save_tracker(df: pd.DataFrame) -> None:
    """Lưu tracker ra cả parquet và CSV."""
    df.to_parquet(TRACKER_PATH, index=False)
    df.to_csv(TRACKER_CSV, index=False, sep=';',
              encoding='utf-8-sig')   # utf-8-sig cho Excel VN


def generate_weekly_report(df: pd.DataFrame) -> dict:
    """
    Tính toán metrics cho weekly Telegram report.
    
    Returns dict với các chỉ số cần thiết.
    """
    df['signal_date'] = pd.to_datetime(df['signal_date'])
    
    # Tổng quan
    total    = len(df)
    holding  = (df['result'] == 'HOLDING').sum()
    wins     = (df['result'] == 'WIN').sum()
    losses   = (df['result'] == 'LOSS').sum()
    closed   = wins + losses
    win_rate = (wins / closed * 100) if closed > 0 else None
    
    # Return trung bình các lệnh đã đóng
    closed_df  = df[df['result'].isin(['WIN', 'LOSS'])].copy()
    avg_return = None
    if len(closed_df) > 0 and 'return_pct' in closed_df.columns:
        avg_return = closed_df['return_pct'].mean()
    
    # Signals trong 7 ngày qua
    week_ago     = pd.Timestamp(date.today() - timedelta(days=7))
    new_signals  = (df['signal_date'] >= week_ago).sum()
    
    # Regime distribution gần đây
    regime_path = PROC_DIR / 'market_regime.parquet'
    regime_dist = {}
    if regime_path.exists():
        r = pd.read_parquet(regime_path).tail(20)
        regime_dist = r['regime'].value_counts().to_dict()
    
    return {
        'total'      : total,
        'holding'    : holding,
        'wins'       : wins,
        'losses'     : losses,
        'closed'     : closed,
        'win_rate'   : win_rate,
        'avg_return' : avg_return,
        'new_signals': new_signals,
        'regime_dist': regime_dist,
    }


def format_weekly_telegram(metrics: dict) -> str:
    """Format Telegram message cho weekly review."""
    today    = date.today().strftime('%d/%m/%Y')
    wr       = metrics['win_rate']
    avg_ret  = metrics['avg_return']
    
    # Win rate status
    if wr is None:
        wr_line = "📊 Win Rate: Chưa có lệnh đóng"
    elif wr >= 52:
        wr_line = f"📊 Win Rate: {wr:.1f}% ✅ (mục tiêu ≥52%)"
    else:
        wr_line = f"📊 Win Rate: {wr:.1f}% ⚠️ (dưới mục tiêu)"
    
    # Avg return
    if avg_ret is not None:
        ret_emoji = "✅" if avg_ret > 0 else "❌"
        ret_line  = f"💰 Avg Return: {avg_ret:+.2f}% {ret_emoji}"
    else:
        ret_line  = "💰 Avg Return: Chưa có dữ liệu"
    
    # Regime summary
    regime_lines = []
    for regime, count in metrics['regime_dist'].items():
        emoji = {'BULL':'🟢','SIDEWAY':'🟡','BEAR':'🔴'}.get(regime,'⚪')
        regime_lines.append(f"  {emoji} {regime}: {count} ngày")
    
    return "\n".join([
        f"📅 <b>WEEKLY REVIEW — {today}</b>",
        "━━━━━━━━━━━━━━━━━━━━",
        f"📈 Signals tuần này : {metrics['new_signals']}",
        f"⏳ Đang HOLDING     : {metrics['holding']}",
        f"✅ WIN              : {metrics['wins']}",
        f"❌ LOSS             : {metrics['losses']}",
        "━━━━━━━━━━━━━━━━━━━━",
        wr_line,
        ret_line,
        "━━━━━━━━━━━━━━━━━━━━",
        "🌐 Regime 20 ngày gần nhất:",
        *regime_lines,
        "━━━━━━━━━━━━━━━━━━━━",
        "📝 <i>Mở live_signal_tracker.csv",
        "để xem chi tiết từng lệnh</i>",
    ])


def sync_to_excel(
    csv_path: Path,
    xlsx_path: Path,
) -> None:
    """
    Sync live_signal_tracker.csv → journal/trade_log.xlsx.
    Preserve cột 'note' do user nhập tay.

    Parameters
    ----------
    csv_path  : Path đến live_signal_tracker.csv (source)
    xlsx_path : Path đến journal/trade_log.xlsx (destination)
    """
    import pandas as pd

    if not csv_path.exists():
        print(f"⚠️  Không tìm thấy CSV source: {csv_path}")
        return

    # 1. Đọc dữ liệu mới nhất từ CSV (auto-detect delimiter ',', ';', '\t') hoặc parquet
    try:
        df_new = pd.read_csv(csv_path, sep=None, engine='python', encoding='utf-8-sig')
    except Exception:
        df_new = pd.DataFrame()

    if 'signal_date' not in df_new.columns:
        parquet_path = csv_path.with_suffix('.parquet')
        if parquet_path.exists():
            df_new = pd.read_parquet(parquet_path)

    if df_new.empty or 'signal_date' not in df_new.columns:
        print(f"⚠️  Không đọc được dữ liệu hợp lệ từ {csv_path}")
        return

    df_new['signal_date'] = pd.to_datetime(df_new['signal_date'])


    # 2. Đọc notes hiện có từ xlsx (nếu file tồn tại)
    existing_notes = {}
    if xlsx_path.exists():
        try:
            df_old = pd.read_excel(xlsx_path, engine='openpyxl')
            if 'signal_date' in df_old.columns:
                df_old['signal_date'] = pd.to_datetime(
                    df_old['signal_date']
                )
            # Key: (signal_date, ticker) → note
            # Chỉ lưu các note không rỗng
            for _, row in df_old.iterrows():
                note = row.get('note', '')
                if pd.notna(note) and str(note).strip():
                    sig_dt = pd.to_datetime(row['signal_date']).date() if pd.notna(row.get('signal_date')) else str(row.get('signal_date'))
                    key = (sig_dt, str(row.get('ticker', '')).strip())
                    existing_notes[key] = str(note).strip()
        except Exception as e:
            print(f"⚠️  Không đọc được xlsx cũ: {e}")

    # 3. Đảm bảo cột 'note' tồn tại trong df_new
    if 'note' not in df_new.columns:
        df_new['note'] = ''

    # 4. Restore notes — không ghi đè nếu user đã điền
    def restore_note(row):
        sig_dt = pd.to_datetime(row['signal_date']).date() if pd.notna(row.get('signal_date')) else str(row.get('signal_date'))
        key = (sig_dt, str(row.get('ticker', '')).strip())
        # Ưu tiên note cũ của user
        if key in existing_notes:
            return existing_notes[key]
        # Giữ nguyên note trong CSV nếu có
        current = row.get('note', '')
        if pd.notna(current) and str(current).strip():
            return str(current).strip()
        return ''

    df_new['note'] = df_new.apply(restore_note, axis=1)

    # 5. Sắp xếp: mới nhất lên đầu
    df_new = df_new.sort_values(
        ['signal_date', 'ticker'],
        ascending=[False, True]
    ).reset_index(drop=True)

    # 6. Format cột số cho dễ đọc
    currency_cols = [
        'entry_price', 'STOP_LOSS', 'TARGET_1',
        'TARGET_2', 'current_price', 'exit_price'
    ]
    for col in currency_cols:
        if col in df_new.columns:
            df_new[col] = pd.to_numeric(
                df_new[col], errors='coerce'
            )

    pct_cols = ['return_pct']
    for col in pct_cols:
        if col in df_new.columns:
            df_new[col] = pd.to_numeric(
                df_new[col], errors='coerce'
            )

    # 7. Ghi ra xlsx với formatting
    _write_xlsx(df_new, xlsx_path)
    print(f"✅ trade_log.xlsx updated: {len(df_new)} signals")


def _write_xlsx(df: pd.DataFrame, xlsx_path: Path) -> None:
    """
    Ghi DataFrame ra xlsx với conditional formatting.
    WIN  → nền xanh lá  (#C6EFCE)
    LOSS → nền đỏ nhạt  (#FFC7CE)
    HOLDING → nền vàng  (#FFEB9C)
    EXPIRED → nền xám   (#E0E0E0)
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment
    )
    from openpyxl.utils import get_column_letter

    # Màu sắc
    COLOR_MAP = {
        'WIN'    : 'C6EFCE',   # xanh lá nhạt
        'LOSS'   : 'FFC7CE',   # đỏ nhạt
        'HOLDING': 'FFEB9C',   # vàng nhạt
        'EXPIRED': 'E0E0E0',   # xám nhạt
    }
    HEADER_COLOR = '1A237E'    # xanh đậm (giống email report)

    wb = Workbook()
    ws = wb.active
    ws.title = "Signal Tracker"

    # ── Header row ──────────────────────────────────────────────
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = Font(
            bold=True, color='FFFFFF', size=11
        )
        cell.fill      = PatternFill(
            fill_type='solid', fgColor=HEADER_COLOR
        )
        cell.alignment = Alignment(
            horizontal='center', vertical='center'
        )

    # ── Data rows ────────────────────────────────────────────────
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        result = str(getattr(row, 'result', '') or '')
        bg_color = COLOR_MAP.get(result, 'FFFFFF')
        fill = PatternFill(fill_type='solid', fgColor=bg_color)

        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # Format giá trị
            col_name = headers[col_idx - 1]

            if col_name in ['entry_price', 'STOP_LOSS',
                            'TARGET_1', 'TARGET_2',
                            'current_price', 'exit_price']:
                if pd.notna(value) and value != '':
                    cell.value  = float(value)
                    cell.number_format = '#,##0'
                else:
                    cell.value = ''

            elif col_name == 'return_pct':
                if pd.notna(value) and value != '':
                    cell.value  = float(value) / 100
                    cell.number_format = '0.00%'
                else:
                    cell.value = ''

            elif col_name in ['signal_date',
                              'last_updated', 'exit_date']:
                cell.value = str(value)[:10] if pd.notna(value) \
                             else ''

            else:
                cell.value = value if pd.notna(value) else ''

            cell.fill      = fill
            cell.alignment = Alignment(
                horizontal='center', vertical='center',
                wrap_text=(col_name == 'note')
            )

    # ── Column widths ────────────────────────────────────────────
    col_widths = {
        'signal_date'  : 14,
        'ticker'       : 8,
        'sector'       : 22,
        'signals_fired': 20,
        'score'        : 8,
        'entry_price'  : 14,
        'STOP_LOSS'    : 14,
        'TARGET_1'     : 14,
        'TARGET_2'     : 14,
        'current_price': 14,
        'return_pct'   : 12,
        'status'       : 12,
        'last_updated' : 14,
        'result'       : 12,
        'days_held'    : 12,
        'exit_price'   : 14,
        'exit_date'    : 14,
        'note'         : 45,   # rộng hơn để ghi chú
    }
    for col_idx, col_name in enumerate(headers, 1):
        width = col_widths.get(col_name, 14)
        ws.column_dimensions[
            get_column_letter(col_idx)
        ].width = width

    # ── Freeze header row ────────────────────────────────────────
    ws.freeze_panes = 'A2'

    # ── Auto filter ──────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    # Đảm bảo thư mục tồn tại
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)

