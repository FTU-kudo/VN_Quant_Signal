import sys
sys.path.insert(0, '.')
import pandas as pd
from pathlib import Path
from config.settings import PROC_DIR
from src.tracking.auto_updater import sync_to_excel

sys.stdout.reconfigure(encoding='utf-8', errors='ignore')

csv_file = PROC_DIR / 'live_signal_tracker.csv'
xlsx_file = Path('journal/trade_log.xlsx')

print("--- RUN 1: SYNC FROM CSV TO XLSX ---")
sync_to_excel(csv_file, xlsx_file)

assert xlsx_file.exists(), "Excel file was not created!"
df1 = pd.read_excel(xlsx_file, engine='openpyxl')
print(f"Rows created in Excel: {len(df1)}")
print(df1[['signal_date', 'ticker', 'entry_price', 'current_price', 'result', 'note']].head(10))

# Test preserving user note
print("\n--- RUN 2: SIMULATE USER ADDING MANUAL NOTE TO KLB ---")
from openpyxl import load_workbook
wb = load_workbook(xlsx_file)
ws = wb.active

# Find header column index for 'note' and 'ticker'
header = [cell.value for cell in ws[1]]
ticker_idx = header.index('ticker') + 1
note_idx = header.index('note') + 1

found_klb = False
for row in range(2, ws.max_row + 1):
    if ws.cell(row=row, column=ticker_idx).value == 'KLB':
        ws.cell(row=row, column=note_idx).value = 'Ghi chú test của User — KHÔNG ĐƯỢC XÓA!'
        found_klb = True
        break

wb.save(xlsx_file)
wb.close()

if found_klb:
    print("Đã nhập tay note vào mã KLB trong trade_log.xlsx.")

print("\n--- RUN 3: RE-SYNC AND VERIFY NOTE IS PRESERVED ---")
sync_to_excel(csv_file, xlsx_file)

df3 = pd.read_excel(xlsx_file, engine='openpyxl')
klb_row = df3[df3['ticker'] == 'KLB']
if not klb_row.empty:
    preserved_note = klb_row.iloc[0]['note']
    print(f"Note của KLB sau khi sync lại: '{preserved_note}'")
    assert 'KHÔNG ĐƯỢC XÓA' in str(preserved_note), f"Lỗi: Note bị mất/ghi đè! Nhận được: {preserved_note}"
    print("✅ XÁC NHẬN: Cột 'note' do user nhập tay được BẢO TỒN 100%!")
else:
    print("Không tìm thấy hàng KLB để kiểm chứng.")
